import os
import requests
import pandas as pd
from supabase import create_client, Client
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from io import BytesIO
from PIL import Image as PILImage
from datetime import datetime
from dotenv import load_dotenv

# Cargar variables de entorno
load_dotenv()

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY") # Debe ser la service_role key para subir archivos si no hay RLS abierto

if not SUPABASE_URL or not SUPABASE_KEY:
    print("Error: SUPABASE_URL o SUPABASE_KEY no definidos en el entorno.")
    exit(1)

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

def generate_client_excel():
    print(f"[{datetime.now()}] Iniciando generación de Excel para Clientes...")

    # 1. Obtener productos de la base de datos
    try:
        response = supabase.table("productos").select("*").execute()
        all_products = response.data
        # ORDENAR por SKU de menor a mayor
        all_products.sort(key=lambda x: int(x.get('codigo') or 0))
    except Exception as e:
        print(f"Error al consultar Supabase: {e}")
        return

    # 2. Libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "LISTA DE PRECIOS"
    headers = ["FOTO", "MARCA", "CÓDIGO", "PRODUCTO", "RUBRO", "DESCRIPCIÓN", "PRECIO PESOS", "PVP", "IVA", "ESTADO"]
    ws.append(headers)

    # Estilos
    header_fill = PatternFill(start_color="4A7CFF", end_color="4A7CFF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    center_aligned = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_aligned

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 8
    ws.column_dimensions['J'].width = 18

    # 3. Procesamiento
    added_count = 0
    current_row = 2
    
    for p in all_products:
        sku = p.get('codigo')
        stock_fisico = float(p.get('stock_fisico') or 0)
        stock_ingresos = float(p.get('stock_ingresos') or 0)
        stock_estado = str(p.get('stock_estado') or '').upper()
        precio = float(p.get('precio_pesos') or 0)
        
        # Filtro de STOCK: Mismas reglas que el catálogo web
        # - Solo productos con stock físico > 1 (o preventa/liquidación)
        # - Debe tener precio > 0
        is_preventa = 'PREVENTA' in stock_estado
        is_liquidacion = 'LIQUIDACION' in stock_estado
        if not is_preventa and not is_liquidacion and stock_fisico <= 1:
            continue
        if stock_estado == 'NOSTOCK':
            continue
        if precio <= 0:
            continue

        # Intentar cargar la foto PRIMERO
        img_url = f"{SUPABASE_URL}/storage/v1/object/public/fotos/{sku}.webp"
        xl_img = None
        try:
            img_res = requests.get(img_url, timeout=10)
            if img_res.status_code == 200:
                pil_img = PILImage.open(BytesIO(img_res.content))
                if pil_img.mode in ("RGBA", "P"):
                    pil_img = pil_img.convert("RGB")
                
                # Redimensionar para que quepa bien en la celda
                pil_img.thumbnail((140, 140))
                img_temp = BytesIO()
                pil_img.save(img_temp, format="JPEG", quality=85)
                img_temp.seek(0)
                xl_img = OpenpyxlImage(img_temp)
            else:
                # FILTRO DE FOTO: Si no tiene foto, el usuario pidió que NO figure
                continue
        except Exception:
            continue # Si hay error de red, saltar este producto

        # Si llegamos aquí, pasó los filtros de stock y foto. Agregamos al excel.
        ws.cell(row=current_row, column=2, value=p.get('marca'))
        ws.cell(row=current_row, column=3, value=sku).font = Font(bold=True)
        ws.cell(row=current_row, column=4, value=p.get('producto'))
        ws.cell(row=current_row, column=5, value=p.get('rubro'))
        ws.cell(row=current_row, column=6, value=p.get('descripcion'))
        
        # Precio con parche para 11573 si fuera necesario (consistencia con web)
        precio_v = float(p.get('precio_pesos') or 0)
        if str(sku) == '11573': precio_v = 59999

        c_price = ws.cell(row=current_row, column=7, value=precio_v)
        c_price.number_format = '"$"#,##0'
        c_pvp = ws.cell(row=current_row, column=8, value=float(p.get('precio_pvp') or 0))
        c_pvp.number_format = '"$"#,##0'
        
        ws.cell(row=current_row, column=9, value=f"{p.get('iva')}%")
        ws.cell(row=current_row, column=10, value=p.get('stock_estado'))

        # Estilo de fila
        ws.row_dimensions[current_row].height = 110
        for col in range(1, 11):
            ws.cell(row=current_row, column=col).alignment = center_aligned

        # Insertar imagen anclada y un poco desplazada para centrar visualmente
        ws.add_image(xl_img, f'A{current_row}')
        
        current_row += 1
        added_count += 1

    # 4. Guardar y Subir
    output_file = "Beepex.xlsx"
    wb.save(output_file)
    print(f"Excel listo con {added_count} productos.")

    try:
        # Intentamos borrar el archivo anterior por si el upsert falla
        try:
            supabase.storage.from_("reportes").remove(["Beepex.xlsx"])
        except Exception:
            pass

        with open(output_file, "rb") as f:
            supabase.storage.from_("reportes").upload(
                path="Beepex.xlsx", 
                file=f, 
                file_options={"x-upsert": "true", "content-type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
            )
        print("Subido a Supabase Storage.")
    except Exception as e:
        print(f"Error subida: {e}")

if __name__ == "__main__":
    generate_client_excel()
